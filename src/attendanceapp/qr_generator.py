"""
Dr. Alfredo Pio De Roda ES - QR Code Generator
Version: IMPROVED v4 - Complete Fix & Better Logic
Date: January 30, 2026

IMPROVEMENTS:
1. ✅ Fixed pack() geometry manager errors
2. ✅ Better UI layout logic
3. ✅ Improved responsiveness
4. ✅ Better error handling
5. ✅ Cleaner code structure
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import load_workbook
import qrcode
import os
from datetime import datetime

class QRGenerator:
    def __init__(self, root):
        self.root = root
        self.root.title("QR Code Generator - Dr. Alfredo Pio De Roda ES")
        self.root.geometry("1000x750")
        self.root.configure(bg="#0f1419")
        
        # Colors
        self.BG_DARK = "#0f1419"
        self.BG_CARD = "#1a202c"
        self.BG_INPUT = "#2d3748"
        self.BLUE = "#3b82f6"
        self.GREEN = "#10b981"
        self.RED = "#ef4444"
        self.YELLOW = "#f59e0b"
        self.CYAN = "#06b6d4"
        self.PURPLE = "#8b5cf6"
        self.TEXT_PRIMARY = "#ffffff"
        self.TEXT_SECONDARY = "#9ca3af"
        
        # Setup folders
        self.home_dir = os.path.expanduser("~")
        self.base_folder = os.path.join(self.home_dir, "SF2_Files")
        self.qr_folder = os.path.join(self.base_folder, "QR_Codes")
        self.active_folder = os.path.join(self.base_folder, "Active")
        
        for folder in [self.qr_folder, self.active_folder]:
            os.makedirs(folder, exist_ok=True)
        
        self.sf2_file = None
        self.student_names = []
        
        self.setup_ui()
    
    def is_valid_student_name(self, name):
        """Validate student name with comprehensive filtering"""
        if not name or not isinstance(name, str):
            return False
        
        name = name.strip()
        if len(name) < 2:
            return False
        
        name_upper = name.upper()
        
        excluded_patterns = [
            "SUMIF", "COUNTIF", "AVERAGE", "SUM(", "COUNT(", "IF(",
            "VLOOKUP", "HLOOKUP", "INDEX", "MATCH",
            "SCHOOL FORM", "SF2", "DAILY ATTENDANCE", "ATTENDANCE REPORT",
            "LEARNER'S NAME", "LAST NAME", "FIRST NAME", "MIDDLE NAME",
            "CODES FOR CHECKING", "PRESENT", "ABSENT", "TARDY",
            "HALF SHADED", "UPPER", "LOWER", "CUTTING CLASSES", "LATE COMER",
            "DROPPED", "TRANSFERRED", "ENROLLED", "REGISTRATION",
            "TOTAL", "COMBINED", "PER DAY", "SUMMARY", "MALE", "FEMALE",
            "MONTH:", "BLANK", "(BLANK)", "NO. OF DAYS", "CLASSES",
            "PERCENTAGE", "ENROLMENT", "AVERAGE DAILY", "ATTENDANCE",
            "REGISTERED LEARNERS", "END OF THE MONTH", "SCHOOL YEAR",
            "1ST FRIDAY", "REPORTING MONTH", "SCHOOL DAYS",
            "REASONS", "CAUSES", "DROPPING OUT", "DROP OUT", "DROPOUT",
            "DOMESTIC-RELATED", "INDIVIDUAL-RELATED", "SCHOOL-RELATED",
            "GEOGRAPHIC", "ENVIRONMENTAL", "FINANCIAL-RELATED",
            "TAKE CARE", "SIBLINGS", "EARLY MARRIAGE", "PREGNANCY",
            "PARENTS' ATTITUDE", "FAMILY PROBLEMS", "ILLNESS",
            "OVERAGE", "DEATH", "DRUG ABUSE", "ACADEMIC PERFORMANCE",
            "LACK OF INTEREST", "DISTRACTIONS", "HUNGER", "MALNUTRITION",
            "TEACHER FACTOR", "PHYSICAL CONDITION", "CLASSROOM",
            "PEER INFLUENCE", "DISTANCE", "HOME AND SCHOOL",
            "ARMED CONFLICT", "TRIBAL WARS", "CLAN FEUDS",
            "CALAMITIES", "DISASTERS", "CHILD LABOR", "WORK",
            "OTHERS (SPECIFY)",
            "GUIDELINES:", "ACCOMPLISHED", "REFER", "DATES SHALL",
            "WRITTEN IN", "COLUMNS AFTER", "COMPUTE", "FOLLOWING",
            "EVERY END", "ADVISER", "SUBMIT", "OFFICE", "PRINCIPAL",
            "RECORDING", "SUMMARY TABLE", "FORM 4", "SIGNED",
            "RETURNED", "PROVIDE", "NECESSARY", "INTERVENTIONS",
            "HOME VISITATION", "ABSENT FOR 5", "CONSECUTIVE DAYS",
            "RISK OF", "PERFORMANCE", "REFLECTED", "FORM 137", "FORM 138",
            "GRADING PERIOD", "BEGINNING", "CUT-OFF",
            "MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY",
            "MONDAY,", "TUESDAY,", "WEDNESDAY,", "THURSDAY,", "FRIDAY,",
            "CERTIFY", "TRUE", "CORRECT", "REPORT", "SIGNATURE",
            "PRINTED NAME", "TEACHER", "SCHOOL HEAD", "ATTESTED",
            "PAGE", "OF", "SCHOOL FORM 2", "___",
            "LEARNER", "STUDENT", "NAME", "NAMES", "ID", "NUMBER",
            "ENROLLMENT", "ENROL",
            "NAN", "NONE", "N/A", "NULL", "BLANK", "EMPTY",
        ]
        
        for pattern in excluded_patterns:
            if pattern in name_upper:
                return False
        
        if not any(c.isalpha() for c in name):
            return False
        
        if name.replace('.', '').replace(',', '').replace(' ', '').isdigit():
            return False
        
        return True
    
    def setup_ui(self):
        """Setup UI with better layout"""
        # Main container
        main_frame = tk.Frame(self.root, bg=self.BG_DARK)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
        
        # Header
        header = tk.Frame(main_frame, bg=self.BG_CARD, relief=tk.RIDGE, bd=2)
        header.pack(fill=tk.X, pady=(0, 15))
        
        tk.Label(header, text="🎯 QR Code Generator", font=("Segoe UI", 16, "bold"),
                fg=self.BLUE, bg=self.BG_CARD).pack(anchor="w", padx=20, pady=(15, 5))
        
        tk.Label(header, text="Generate QR codes for all students in your SF2 file", 
                font=("Segoe UI", 10), fg=self.TEXT_SECONDARY, bg=self.BG_CARD).pack(anchor="w", padx=20, pady=(0, 15))
        
        # File selection section
        file_section = tk.Frame(main_frame, bg=self.BG_CARD, relief=tk.RIDGE, bd=2)
        file_section.pack(fill=tk.X, pady=(0, 15))
        
        tk.Label(file_section, text="📄 Select SF2 File", font=("Segoe UI", 12, "bold"),
                fg=self.GREEN, bg=self.BG_CARD).pack(anchor="w", padx=15, pady=(10, 10))
        
        file_btn_frame = tk.Frame(file_section, bg=self.BG_CARD)
        file_btn_frame.pack(fill=tk.X, padx=15, pady=(0, 15))
        
        tk.Button(file_btn_frame, text="📁 Browse & Load SF2 File", command=self.browse_file,
                 bg=self.BLUE, fg="#fff", font=("Segoe UI", 11, "bold"),
                 relief=tk.FLAT, padx=20, pady=10, cursor="hand2").pack(side=tk.LEFT)
        
        self.file_label = tk.Label(file_btn_frame, text="No file selected",
                                   font=("Segoe UI", 10), fg=self.TEXT_SECONDARY, bg=self.BG_CARD)
        self.file_label.pack(side=tk.LEFT, padx=20)
        
        # Info section
        info_section = tk.Frame(main_frame, bg=self.BG_CARD, relief=tk.RIDGE, bd=2)
        info_section.pack(fill=tk.X, pady=(0, 15))
        
        tk.Label(info_section, text="📊 File Information", font=("Segoe UI", 12, "bold"),
                fg=self.CYAN, bg=self.BG_CARD).pack(anchor="w", padx=15, pady=(10, 5))
        
        self.info_text = tk.Label(info_section, text="Students: 0\nQR Codes: 0\nStatus: Ready to load",
                                 font=("Segoe UI", 10), fg=self.TEXT_PRIMARY, bg=self.BG_CARD)
        self.info_text.pack(anchor="w", padx=15, pady=(0, 15))
        
        # Progress section
        progress_section = tk.Frame(main_frame, bg=self.BG_CARD, relief=tk.RIDGE, bd=2)
        progress_section.pack(fill=tk.X, pady=(0, 15))
        
        tk.Label(progress_section, text="⏳ Generation Progress", font=("Segoe UI", 12, "bold"),
                fg=self.YELLOW, bg=self.BG_CARD).pack(anchor="w", padx=15, pady=(10, 10))
        
        self.progress = ttk.Progressbar(progress_section, mode='determinate', maximum=100)
        self.progress.pack(fill=tk.X, padx=15, pady=(0, 10))
        
        self.status_label = tk.Label(progress_section, text="Status: Idle",
                                    font=("Segoe UI", 10), fg=self.TEXT_SECONDARY, bg=self.BG_CARD)
        self.status_label.pack(anchor="w", padx=15, pady=(0, 15))
        
        # Generate buttons
        gen_btn_frame = tk.Frame(main_frame, bg=self.BG_DARK)
        gen_btn_frame.pack(fill=tk.X, pady=(0, 15))
        
        self.generate_btn = tk.Button(gen_btn_frame, text="🎯 Generate All QR Codes", 
                                     command=self.generate_qr_codes,
                                     bg=self.GREEN, fg="#000", font=("Segoe UI", 12, "bold"),
                                     relief=tk.FLAT, padx=30, pady=12, cursor="hand2",
                                     state=tk.DISABLED)
        self.generate_btn.pack(side=tk.LEFT, padx=(0, 5), fill=tk.X, expand=True)
        
        self.open_folder_btn = tk.Button(gen_btn_frame, text="📂 QR FOLDER", 
                                        command=self.open_qr_folder,
                                        bg=self.PURPLE, fg="#fff", font=("Segoe UI", 12, "bold"),
                                        relief=tk.FLAT, padx=30, pady=12, cursor="hand2")
        self.open_folder_btn.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        
        self.open_output_btn = tk.Button(gen_btn_frame, text="📁 OUTPUT", 
                                        command=self.open_output_folder,
                                        bg=self.YELLOW, fg="#000", font=("Segoe UI", 12, "bold"),
                                        relief=tk.FLAT, padx=30, pady=12, cursor="hand2")
        self.open_output_btn.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Info text
        info_frame = tk.Frame(main_frame, bg=self.BG_CARD, relief=tk.RIDGE, bd=2)
        info_frame.pack(fill=tk.BOTH, expand=True)
        
        tk.Label(info_frame, text="ℹ️ How It Works", font=("Segoe UI", 12, "bold"),
                fg=self.BLUE, bg=self.BG_CARD).pack(anchor="w", padx=15, pady=(10, 5))
        
        help_text = """
1. Click "Browse & Load SF2 File" to select your Excel file
2. System loads all valid student names from Column B
3. Click "Generate All QR Codes" to create QR codes
4. QR codes are saved to: {folder}
5. Print the QR codes and attach to student IDs
6. Use the Attendance System to scan QR codes

WHAT IT GENERATES:
• Individual QR code images for each student
• QR code contains student name
• Organized in QR_Codes folder
• Easy to print and laminate

FILE STRUCTURE:
• Row 13+: Students (Column B = Name)
• Names are validated (no headers or formulas)
• Only real student names get QR codes

TIPS:
• Print on sticker sheets for easy ID attachment
• Use 2"x2" or 1.5"x1.5" size for best scanning
• Laminate for durability
• Organize students by grade level
        """.format(folder=self.qr_folder)
        
        tk.Label(info_frame, text=help_text, font=("Segoe UI", 9),
                fg=self.TEXT_PRIMARY, bg=self.BG_CARD, justify=tk.LEFT).pack(anchor="w", padx=15, pady=(0, 15))
    
    def browse_file(self):
        """Browse for SF2 file"""
        file_path = filedialog.askopenfilename(
            initialdir=self.active_folder,
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if file_path:
            self.load_file(file_path)
    
    def load_file(self, file_path):
        """Load SF2 file"""
        try:
            print(f"\n{'='*80}")
            print(f"LOADING FILE: {os.path.basename(file_path)}")
            print(f"{'='*80}")
            
            workbook = load_workbook(file_path)
            sheet = workbook.active
            
            self.sf2_file = file_path
            self.student_names = []
            
            print(f"👥 Extracting students from Column B...")
            print(f"Using SMART FILTERING (only real students, skips unrelated data)...")
            print("-" * 80)
            
            for row in range(13, sheet.max_row + 1):
                name_cell = sheet.cell(row, 2).value
                
                # Skip empty rows
                if not name_cell:
                    continue
                
                if self.is_valid_student_name(name_cell):
                    name = name_cell.strip()
                    self.student_names.append(name)
                    print(f"  Row {row:3d}: ✅ {name}")
                else:
                    # Log filtered entries (unrelated text)
                    print(f"  Row {row:3d}: ⊘  FILTERED - '{name_cell}'")
            
            print("-" * 80)
            print(f"✅ Loaded {len(self.student_names)} valid students")
            print(f"{'='*80}\n")
            
            # Update UI
            self.file_label.config(text=f"✅ Loaded: {os.path.basename(file_path)}", 
                                  fg=self.GREEN)
            self.info_text.config(
                text=f"Students: {len(self.student_names)}\nQR Codes: Ready to generate\nStatus: File loaded successfully",
                fg=self.GREEN
            )
            self.generate_btn.config(state=tk.NORMAL)
        
        except Exception as e:
            print(f"❌ Error loading file: {e}")
            messagebox.showerror("Error", f"Failed to load file: {e}")
    
    def generate_qr_codes(self):
        """Generate QR codes for all students"""
        if not self.student_names:
            messagebox.showwarning("Warning", "No students loaded!")
            return
        
        try:
            print(f"\n{'='*80}")
            print(f"GENERATING QR CODES")
            print(f"{'='*80}")
            print(f"Total students: {len(self.student_names)}")
            print(f"Output folder: {self.qr_folder}")
            print("-" * 80)
            
            self.generate_btn.config(state=tk.DISABLED)
            self.progress['maximum'] = len(self.student_names)
            
            for index, student_name in enumerate(self.student_names):
                # Generate QR code
                qr = qrcode.QRCode(
                    version=1,
                    error_correction=qrcode.constants.ERROR_CORRECT_H,
                    box_size=10,
                    border=4,
                )
                qr.add_data(student_name)
                qr.make(fit=True)
                
                img = qr.make_image(fill_color="black", back_color="white")
                
                # Save QR code
                filename = os.path.join(self.qr_folder, f"{student_name.replace(' ', '_')}.png")
                img.save(filename)
                
                print(f"  ✅ {index+1:3d}/{len(self.student_names)}: {student_name}")
                
                # Update progress
                self.progress['value'] = index + 1
                self.status_label.config(text=f"Status: Generated {index+1}/{len(self.student_names)}")
                self.root.update()
            
            print("-" * 80)
            print(f"✅ All {len(self.student_names)} QR codes generated successfully!")
            print(f"📁 Location: {self.qr_folder}")
            print(f"{'='*80}\n")
            
            self.status_label.config(text=f"Status: ✅ Completed! Generated {len(self.student_names)} QR codes",
                                    fg=self.GREEN)
            messagebox.showinfo("Success", f"✅ Generated {len(self.student_names)} QR codes successfully!\n\nLocation: {self.qr_folder}")
            
            self.generate_btn.config(state=tk.NORMAL)
        
        except Exception as e:
            print(f"❌ Error generating QR codes: {e}")
            messagebox.showerror("Error", f"Failed to generate QR codes: {e}")
            self.generate_btn.config(state=tk.NORMAL)
    
    def open_qr_folder(self):
        """Open QR folder"""
        try:
            import subprocess
            if os.name == 'nt':  # Windows
                subprocess.Popen(f'explorer "{self.qr_folder}"')
            elif os.name == 'posix':  # Mac/Linux
                subprocess.Popen(['open', self.qr_folder])
        except Exception as e:
            messagebox.showerror("Error", f"Cannot open folder: {e}")
    
    def open_output_folder(self):
        """Open Output folder (where files are saved)"""
        try:
            import subprocess
            output_folder = os.path.join(os.path.expanduser("~"), "Downloads", "ATStudios-Project")
            if os.name == 'nt':  # Windows
                subprocess.Popen(f'explorer "{output_folder}"')
            elif os.name == 'posix':  # Mac/Linux
                subprocess.Popen(['open', output_folder])
        except Exception as e:
            messagebox.showerror("Error", f"Cannot open folder: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = QRGenerator(root)
    root.mainloop()
