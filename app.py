"""
Dr. Alfredo Pio De Roda ES - QR Attendance System
Version: PERFECTED - TABS + FILTERING + EXCEL CHECK + CAMERA FIX
Date: January 30, 2026

PERFECTED:
1. ✅ ORIGINAL UI WITH TABS (exactly like original)
2. ✅ FILTER: Percentage, Average Daily, Attendance percentages
3. ✅ CHECK EXCEL: Existing marks on load
4. ✅ ACCURATE COUNTER: Existing + new scans
5. ✅ SMOOTH CAMERA: Threading for no freezing
6. ✅ AUTO-SAVE: Each scan saved immediately
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import cv2
from PIL import Image, ImageTk
from pyzbar.pyzbar import decode
from openpyxl import load_workbook
from datetime import datetime
import os
import traceback
import re
import threading
import queue
import time
import shutil
from zipfile import ZipFile
import tempfile

class AttendanceSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("QR Attendance System - Dr. Alfredo Pio De Roda ES")
        self.root.geometry("1600x900")
        self.root.configure(bg="#0f1419")
        
        # Initialize variables
        self.camera = None
        self.camera_active = False
        self.camera_thread = None
        self.frame_queue = queue.Queue(maxsize=1)
        self.sf2_workbook = None
        self.sf2_sheet = None
        self.sf2_file = None
        self.student_names = []
        self.scanned_today = []
        self.existing_marks = {}  # Track existing ✓ from Excel
        self.current_column = None
        self.last_scanned = None  # Track last scan to prevent rapid re-scans
        self.last_scan_time = 0  # Track scan time
        
        # Colors
        self.BG_DARK = "#0f1419"
        self.BG_CARD = "#1a202c"
        self.BG_INPUT = "#2d3748"
        self.BLUE = "#3b82f6"
        self.GREEN = "#10b981"
        self.RED = "#ef4444"
        self.YELLOW = "#f59e0b"
        self.CYAN = "#06b6d4"
        self.PURPLE = "#8b5cf6"
        self.TEXT_PRIMARY = "#ffffff"
        self.TEXT_SECONDARY = "#9ca3af"
        
        # Setup folders
        self.home_dir = os.path.expanduser("~")
        self.base_folder = os.path.join(self.home_dir, "SF2_Files")
        self.active_folder = os.path.join(self.base_folder, "Active")
        self.backup_folder = os.path.join(self.base_folder, "Backups")
        self.archive_folder = os.path.join(self.base_folder, "Archive")
        self.qr_folder = os.path.join(self.base_folder, "QR_Codes")
        
        for folder in [self.active_folder, self.backup_folder, self.archive_folder, self.qr_folder]:
            os.makedirs(folder, exist_ok=True)
        
        self.setup_ui()
        self.auto_load_file()
    
    def is_valid_student_name(self, name):
        """Validate if text is a real student name with comprehensive filtering"""
        if not name or not isinstance(name, str):
            return False
        
        name = name.strip()
        if len(name) < 2:
            return False
        
        name_upper = name.upper()
        
        excluded_patterns = [
            "SUMIF", "COUNTIF", "AVERAGE", "SUM(", "COUNT(", "IF(",
            "VLOOKUP", "HLOOKUP", "INDEX", "MATCH",
            "SCHOOL FORM", "SF2", "DAILY ATTENDANCE", "ATTENDANCE REPORT",
            "LEARNER'S NAME", "LAST NAME", "FIRST NAME", "MIDDLE NAME",
            "CODES FOR CHECKING", "PRESENT", "ABSENT", "TARDY",
            "HALF SHADED", "UPPER", "LOWER", "CUTTING CLASSES", "LATE COMER",
            "DROPPED", "TRANSFERRED", "ENROLLED", "REGISTRATION",
            "TOTAL", "COMBINED", "PER DAY", "SUMMARY", "MALE", "FEMALE",
            "MONTH:", "BLANK", "(BLANK)", "NO. OF DAYS", "CLASSES",
            "PERCENTAGE", "ENROLMENT", "AVERAGE DAILY", "ATTENDANCE",
            "REGISTERED LEARNERS", "END OF THE MONTH", "SCHOOL YEAR",
            "1ST FRIDAY", "REPORTING MONTH", "SCHOOL DAYS",
            "REASONS", "CAUSES", "DROPPING OUT", "DROP OUT", "DROPOUT",
            "DOMESTIC-RELATED", "INDIVIDUAL-RELATED", "SCHOOL-RELATED",
            "GEOGRAPHIC", "ENVIRONMENTAL", "FINANCIAL-RELATED",
            "TAKE CARE", "SIBLINGS", "EARLY MARRIAGE", "PREGNANCY",
            "PARENTS' ATTITUDE", "FAMILY PROBLEMS", "ILLNESS",
            "OVERAGE", "DEATH", "DRUG ABUSE", "ACADEMIC PERFORMANCE",
            "LACK OF INTEREST", "DISTRACTIONS", "HUNGER", "MALNUTRITION",
            "TEACHER FACTOR", "PHYSICAL CONDITION", "CLASSROOM",
            "PEER INFLUENCE", "DISTANCE", "HOME AND SCHOOL",
            "ARMED CONFLICT", "TRIBAL WARS", "CLAN FEUDS",
            "CALAMITIES", "DISASTERS", "CHILD LABOR", "WORK",
            "OTHERS (SPECIFY)",
            "GUIDELINES:", "ACCOMPLISHED", "REFER", "DATES SHALL",
            "WRITTEN IN", "COLUMNS AFTER", "COMPUTE", "FOLLOWING",
            "EVERY END", "ADVISER", "SUBMIT", "OFFICE", "PRINCIPAL",
            "RECORDING", "SUMMARY TABLE", "FORM 4", "SIGNED",
            "RETURNED", "PROVIDE", "NECESSARY", "INTERVENTIONS",
            "HOME VISITATION", "ABSENT FOR 5", "CONSECUTIVE DAYS",
            "RISK OF", "PERFORMANCE", "REFLECTED", "FORM 137", "FORM 138",
            "GRADING PERIOD", "BEGINNING", "CUT-OFF",
            "MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY",
            "MONDAY,", "TUESDAY,", "WEDNESDAY,", "THURSDAY,", "FRIDAY,",
            "CERTIFY", "TRUE", "CORRECT", "REPORT", "SIGNATURE",
            "PRINTED NAME", "TEACHER", "SCHOOL HEAD", "ATTESTED",
            "PAGE", "OF", "SCHOOL FORM 2", "___",
            "LEARNER", "STUDENT", "NAME", "NAMES", "ID", "NUMBER",
            "ENROLLMENT", "ENROL",
            "NAN", "NONE", "N/A", "NULL", "BLANK", "EMPTY",
            # ADD NEW FILTERS
            "PERCENTAGE OF ENROLMENT", "PERCENTAGE OF ENROLLMENT",
            "AVERAGE DAILY ATTENDANCE", 
            "PERCENTAGE OF ATTENDANCE FOR THE MONTH",
            "PERCENTAGE OF ATTENDANCE",
        ]
        
        for pattern in excluded_patterns:
            if pattern in name_upper:
                return False
        
        if not any(c.isalpha() for c in name):
            return False
        
        if name.replace('.', '').replace(',', '').replace(' ', '').isdigit():
            return False
        
        if re.match(r'^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}$', name):
            return False
        
        if not any(c.isalnum() for c in name):
            return False
        
        return True
    
    def is_excel_file_open(self, file_path):
        """Check if Excel file is open/locked"""
        try:
            # Try to rename the file - if it's open, this will fail
            temp_name = file_path + ".tmp"
            os.rename(file_path, temp_name)
            os.rename(temp_name, file_path)
            return False  # File is not open
        except (OSError, IOError):
            return True  # File is open/locked
    
    def setup_ui(self):
        """Setup the complete UI with better layout"""
        # Create notebook (tabs)
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Style tabs
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('TNotebook', background=self.BG_DARK, borderwidth=0)
        style.configure('TNotebook.Tab', background=self.BG_CARD, foreground=self.TEXT_PRIMARY,
                       padding=[20, 10], font=("Segoe UI", 10))
        style.map('TNotebook.Tab', background=[("selected", self.BLUE)])
        
        # Create tabs
        self.scan_tab = ttk.Frame(self.notebook)
        self.files_tab = ttk.Frame(self.notebook)
        self.preview_tab = ttk.Frame(self.notebook)
        self.settings_tab = ttk.Frame(self.notebook)
        
        self.notebook.add(self.scan_tab, text="📱 SCAN")
        self.notebook.add(self.files_tab, text="📂 FILES")
        self.notebook.add(self.preview_tab, text="📊 PREVIEW")
        self.notebook.add(self.settings_tab, text="⚙️ SETTINGS")
        
        # Setup each tab
        self.setup_scan_tab()
        self.setup_files_tab()
        self.setup_preview_tab()
        self.setup_settings_tab()
    
    def setup_scan_tab(self):
        """Setup SCAN tab with better layout logic using GRID"""
        main_frame = tk.Frame(self.scan_tab, bg=self.BG_DARK)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Configure grid - LEFT gets more space, RIGHT is fixed
        main_frame.grid_columnconfigure(0, weight=2)
        main_frame.grid_columnconfigure(1, weight=0, minsize=380)
        main_frame.grid_rowconfigure(0, weight=1)
        
        # LEFT SIDE - Camera
        left_frame = tk.Frame(main_frame, bg=self.BG_CARD, relief=tk.RIDGE, bd=2)
        left_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 10), pady=0)
        
        tk.Label(left_frame, text="📷 Camera Feed", font=("Segoe UI", 12, "bold"),
                fg=self.BLUE, bg=self.BG_CARD).pack(anchor="w", padx=15, pady=(10, 5))
        
        self.camera_label = tk.Label(left_frame, bg=self.BG_INPUT, width=500, height=400)
        self.camera_label.pack(padx=15, pady=(0, 15), fill=tk.BOTH, expand=True)
        
        # RIGHT SIDE - Controls
        right_frame = tk.Frame(main_frame, bg=self.BG_DARK)
        right_frame.grid(row=0, column=1, sticky="nsew", padx=0, pady=0)
        
        # Status card
        status_card = tk.Frame(right_frame, bg=self.BG_CARD, relief=tk.RIDGE, bd=2)
        status_card.pack(fill=tk.X, pady=(0, 10))
        
        tk.Label(status_card, text="📊 System Status", font=("Segoe UI", 12, "bold"),
                fg=self.CYAN, bg=self.BG_CARD).pack(anchor="w", padx=15, pady=(10, 5))
        
        self.file_status = tk.Label(status_card, text="📄 File: Not loaded",
                                    font=("Segoe UI", 9), fg=self.TEXT_SECONDARY, bg=self.BG_CARD)
        self.file_status.pack(anchor="w", padx=15, pady=2)
        
        self.date_status = tk.Label(status_card, text="📅 Date: Not detected",
                                    font=("Segoe UI", 9), fg=self.TEXT_SECONDARY, bg=self.BG_CARD)
        self.date_status.pack(anchor="w", padx=15, pady=2)
        
        self.student_count_label = tk.Label(status_card, text="👥 Students: 0",
                                           font=("Segoe UI", 9), fg=self.TEXT_SECONDARY, bg=self.BG_CARD)
        self.student_count_label.pack(anchor="w", padx=15, pady=(2, 10))
        
        # Counters
        counters_frame = tk.Frame(right_frame, bg=self.BG_DARK)
        counters_frame.pack(fill=tk.X, pady=(0, 10))
        
        present_card = tk.Frame(counters_frame, bg=self.GREEN, relief=tk.RAISED, bd=2)
        present_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        
        tk.Label(present_card, text="✅ Present", font=("Segoe UI", 9, "bold"),
                fg="#000", bg=self.GREEN).pack(pady=(5, 0))
        self.present_label = tk.Label(present_card, text="0", font=("Segoe UI", 18, "bold"),
                                     fg="#000", bg=self.GREEN)
        self.present_label.pack(pady=(0, 5))
        
        absent_card = tk.Frame(counters_frame, bg=self.RED, relief=tk.RAISED, bd=2)
        absent_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
        
        tk.Label(absent_card, text="❌ Absent", font=("Segoe UI", 9, "bold"),
                fg="#fff", bg=self.RED).pack(pady=(5, 0))
        self.absent_label = tk.Label(absent_card, text="0", font=("Segoe UI", 18, "bold"),
                                    fg="#fff", bg=self.RED)
        self.absent_label.pack(pady=(0, 5))
        
        total_card = tk.Frame(counters_frame, bg=self.BLUE, relief=tk.RAISED, bd=2)
        total_card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 0))
        
        tk.Label(total_card, text="👥 Total", font=("Segoe UI", 9, "bold"),
                fg="#fff", bg=self.BLUE).pack(pady=(5, 0))
        self.total_label = tk.Label(total_card, text="0", font=("Segoe UI", 18, "bold"),
                                   fg="#fff", bg=self.BLUE)
        self.total_label.pack(pady=(0, 5))
        
        # Scanned list
        list_frame = tk.Frame(right_frame, bg=self.BG_CARD, relief=tk.RIDGE, bd=2)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        tk.Label(list_frame, text="✅ Scanned", font=("Segoe UI", 11, "bold"),
                fg=self.GREEN, bg=self.BG_CARD).pack(anchor="w", padx=15, pady=(10, 5))
        
        table_container = tk.Frame(list_frame, bg=self.BG_CARD)
        table_container.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 10))
        
        columns = ("Name", "Time")
        self.student_tree = ttk.Treeview(table_container, columns=columns, show="headings", height=10)
        
        self.student_tree.heading("Name", text="Student Name")
        self.student_tree.heading("Time", text="Time")
        
        self.student_tree.column("Name", width=200)
        self.student_tree.column("Time", width=80)
        
        scrollbar = ttk.Scrollbar(table_container, orient=tk.VERTICAL, command=self.student_tree.yview)
        self.student_tree.configure(yscrollcommand=scrollbar.set)
        
        self.student_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Style treeview
        style = ttk.Style()
        style.configure("Treeview", background=self.BG_INPUT, foreground=self.TEXT_PRIMARY,
                       fieldbackground=self.BG_INPUT, font=("Segoe UI", 9))
        style.configure("Treeview.Heading", background=self.BG_CARD, foreground=self.GREEN,
                       font=("Segoe UI", 9, "bold"))
        
        # Buttons
        button_frame = tk.Frame(right_frame, bg=self.BG_DARK)
        button_frame.pack(fill=tk.X)
        
        self.start_btn = tk.Button(button_frame, text="▶ START SCANNING", command=self.start_camera,
                                   bg=self.GREEN, fg="#000", font=("Segoe UI", 10, "bold"),
                                   relief=tk.FLAT, padx=15, pady=8, cursor="hand2")
        self.start_btn.pack(fill=tk.X, pady=(0, 4))
        
        self.stop_btn = tk.Button(button_frame, text="⏹ STOP SCANNING", command=self.stop_camera,
                                  bg=self.RED, fg="#fff", font=("Segoe UI", 10, "bold"),
                                  relief=tk.FLAT, padx=15, pady=8, cursor="hand2", state=tk.DISABLED)
        self.stop_btn.pack(fill=tk.X, pady=(0, 4))
        
        tk.Label(button_frame, text="♻️ Continuous Auto-Scanning", font=("Segoe UI", 9, "bold"),
                fg=self.GREEN, bg=self.BG_DARK).pack(fill=tk.X, pady=(8, 4))
        
        self.open_qr_btn = tk.Button(button_frame, text="📂 QR FOLDER", command=self.open_qr_folder,
                                     bg=self.PURPLE, fg="#fff", font=("Segoe UI", 10, "bold"),
                                     relief=tk.FLAT, padx=15, pady=8, cursor="hand2")
        self.open_qr_btn.pack(fill=tk.X, pady=(0, 4))
        
        self.open_active_btn = tk.Button(button_frame, text="📊 ACTIVE FOLDER", command=self.open_active_folder,
                                         bg=self.CYAN, fg="#000", font=("Segoe UI", 10, "bold"),
                                         relief=tk.FLAT, padx=15, pady=8, cursor="hand2")
        self.open_active_btn.pack(fill=tk.X, pady=(0, 4))
        
        self.open_output_btn = tk.Button(button_frame, text="📁 OUTPUT FOLDER", command=self.open_output_folder,
                                         bg=self.YELLOW, fg="#000", font=("Segoe UI", 10, "bold"),
                                         relief=tk.FLAT, padx=15, pady=8, cursor="hand2")
        self.open_output_btn.pack(fill=tk.X)
    
    def setup_files_tab(self):
        """Setup FILES tab"""
        main_frame = tk.Frame(self.files_tab, bg=self.BG_DARK)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        header = tk.Frame(main_frame, bg=self.BG_CARD, relief=tk.RIDGE, bd=2)
        header.pack(fill=tk.X, pady=(0, 15))
        
        tk.Label(header, text="📂 File Manager", font=("Segoe UI", 14, "bold"),
                fg=self.BLUE, bg=self.BG_CARD).pack(anchor="w", padx=15, pady=10)
        
        btn_frame = tk.Frame(main_frame, bg=self.BG_DARK)
        btn_frame.pack(fill=tk.X, pady=(0, 15))
        
        tk.Button(btn_frame, text="🔄 Refresh", command=self.refresh_file_list,
                 bg=self.BLUE, fg="#fff", font=("Segoe UI", 10, "bold"),
                 relief=tk.FLAT, padx=20, pady=10, cursor="hand2").pack(side=tk.LEFT, padx=(0, 10))
        
        tk.Button(btn_frame, text="📁 Browse", command=self.browse_file,
                 bg=self.GREEN, fg="#000", font=("Segoe UI", 10, "bold"),
                 relief=tk.FLAT, padx=20, pady=10, cursor="hand2").pack(side=tk.LEFT)
        
        list_frame = tk.Frame(main_frame, bg=self.BG_CARD, relief=tk.RIDGE, bd=2)
        list_frame.pack(fill=tk.BOTH, expand=True)
        
        tk.Label(list_frame, text="📄 Files in Active Folder", font=("Segoe UI", 12, "bold"),
                fg=self.YELLOW, bg=self.BG_CARD).pack(anchor="w", padx=15, pady=10)
        
        tree_container = tk.Frame(list_frame, bg=self.BG_CARD)
        tree_container.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 15))
        
        columns = ("Filename", "Size", "Modified")
        self.file_tree = ttk.Treeview(tree_container, columns=columns, show="headings", height=20)
        
        self.file_tree.heading("Filename", text="Filename")
        self.file_tree.heading("Size", text="Size")
        self.file_tree.heading("Modified", text="Modified")
        
        self.file_tree.column("Filename", width=400)
        self.file_tree.column("Size", width=100)
        self.file_tree.column("Modified", width=200)
        
        scrollbar = ttk.Scrollbar(tree_container, orient=tk.VERTICAL, command=self.file_tree.yview)
        self.file_tree.configure(yscrollcommand=scrollbar.set)
        
        self.file_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.refresh_file_list()
    
    def setup_preview_tab(self):
        """Setup PREVIEW tab"""
        main_frame = tk.Frame(self.preview_tab, bg=self.BG_DARK)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        header = tk.Frame(main_frame, bg=self.BG_CARD, relief=tk.RIDGE, bd=2)
        header.pack(fill=tk.X, pady=(0, 15))
        
        tk.Label(header, text="👥 Student List", font=("Segoe UI", 14, "bold"),
                fg=self.GREEN, bg=self.BG_CARD).pack(anchor="w", padx=15, pady=10)
        
        tree_frame = tk.Frame(main_frame, bg=self.BG_CARD, relief=tk.RIDGE, bd=2)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        columns = ("Number", "Name", "Status")
        self.preview_tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=30)
        
        self.preview_tree.heading("Number", text="ID")
        self.preview_tree.heading("Name", text="Student Name")
        self.preview_tree.heading("Status", text="Status")
        
        self.preview_tree.column("Number", width=80)
        self.preview_tree.column("Name", width=350)
        self.preview_tree.column("Status", width=100)
        
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.preview_tree.yview)
        self.preview_tree.configure(yscrollcommand=scrollbar.set)
        
        self.preview_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=15, pady=15)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 15), pady=15)
    
    def setup_settings_tab(self):
        """Setup SETTINGS tab"""
        main_frame = tk.Frame(self.settings_tab, bg=self.BG_DARK)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        header = tk.Frame(main_frame, bg=self.BG_CARD, relief=tk.RIDGE, bd=2)
        header.pack(fill=tk.X, pady=(0, 15))
        
        tk.Label(header, text="⚙️ Settings & Information", font=("Segoe UI", 14, "bold"),
                fg=self.YELLOW, bg=self.BG_CARD).pack(anchor="w", padx=15, pady=10)
        
        folder_frame = tk.Frame(main_frame, bg=self.BG_CARD, relief=tk.RIDGE, bd=2)
        folder_frame.pack(fill=tk.X, pady=(0, 15))
        
        tk.Label(folder_frame, text="📁 Folder Locations", font=("Segoe UI", 12, "bold"),
                fg=self.BLUE, bg=self.BG_CARD).pack(anchor="w", padx=15, pady=10)
        
        tk.Label(folder_frame, text=f"Active: {self.active_folder}",
                font=("Segoe UI", 9), fg=self.TEXT_PRIMARY, bg=self.BG_CARD).pack(anchor="w", padx=15, pady=2)
        tk.Label(folder_frame, text=f"QR Codes: {self.qr_folder}",
                font=("Segoe UI", 9), fg=self.TEXT_PRIMARY, bg=self.BG_CARD).pack(anchor="w", padx=15, pady=2)
        
        info_frame = tk.Frame(main_frame, bg=self.BG_CARD, relief=tk.RIDGE, bd=2)
        info_frame.pack(fill=tk.BOTH, expand=True)
        
        tk.Label(info_frame, text="📖 About This System", font=("Segoe UI", 12, "bold"),
                fg=self.GREEN, bg=self.BG_CARD).pack(anchor="w", padx=15, pady=10)
        
        info_text = """
Version: IMPROVED v5 (Simple Logic + Auto-Save)
Date: January 30, 2026

FEATURES:
✅ Simple date column detection (no offset math!)
✅ Auto-save on every scan (no manual save button)
✅ Excel lock detection (warns if file is open)
✅ Real-time counters
✅ QR code scanning
✅ Multiple tabs
✅ Professional dark UI
✅ 100% offline operation

HOW IT WORKS:
1. Find today's date in Row 11
2. That column = attendance column
3. Just put ✓ or x in that column!

DAILY USAGE:
1. File auto-loads from Active folder
2. Click START to begin camera
3. Students scan QR codes
4. Attendance auto-saves (no save button!)
5. If Excel is open, you'll get a warning

SF2 FILE STRUCTURE:
• Row 11: Dates (1, 2, 5, 6, 7...)
• Row 12: Days (M, T, W, TH, F...)
• Row 13+: Students (Col A = ID, Col B = Name)

Auto-Save:
✅ Saves after EACH scan
✅ Shows message if Excel is open
✅ No need to click SAVE!
        """
        
        tk.Label(info_frame, text=info_text, font=("Segoe UI", 9),
                fg=self.TEXT_PRIMARY, bg=self.BG_CARD, justify=tk.LEFT).pack(anchor="w", padx=15, pady=(0, 10))
    
    def auto_load_file(self):
        """Auto-load file from Active folder and start camera"""
        try:
            files = [f for f in os.listdir(self.active_folder) 
                    if f.endswith(('.xlsx', '.xls')) and not f.startswith('~')]
            
            if files:
                file_path = os.path.join(self.active_folder, files[0])
                self.load_file(file_path)
                print(f"✅ Auto-loaded: {files[0]}")
                
                # AUTO-START CAMERA!
                self.root.after(500, self.start_camera)
            else:
                print("⚠️  No Excel files in Active folder")
                self.file_status.config(text="📄 File: No file found", fg=self.RED)
        except Exception as e:
            print(f"❌ Error: {e}")
    
    def repair_excel_file(self, file_path):
        """Attempt to repair corrupted Excel file"""
        try:
            print(f"🔧 Attempting to repair Excel file...")
            backup_path = file_path + ".backup"
            shutil.copy2(file_path, backup_path)
            print(f"  ✓ Backup created")
            
            temp_dir = tempfile.mkdtemp()
            try:
                with ZipFile(file_path, 'r') as zip_ref:
                    zip_ref.extractall(temp_dir)
            except Exception as e:
                return False
            
            content_types_path = os.path.join(temp_dir, '[Content_Types].xml')
            if not os.path.exists(content_types_path):
                return False
            
            try:
                os.remove(file_path)
                with ZipFile(file_path, 'w') as zip_ref:
                    for root, dirs, files in os.walk(temp_dir):
                        for file in files:
                            file_path_full = os.path.join(root, file)
                            arcname = os.path.relpath(file_path_full, temp_dir)
                            zip_ref.write(file_path_full, arcname)
                print(f"  ✓ Repaired successfully!")
                return True
            except Exception as e:
                return False
        except Exception as e:
            return False
    
    def load_file(self, file_path):
        """Load SF2 file with SIMPLE logic + Excel checking"""
        try:
            print(f"\n{'='*80}")
            print(f"LOADING FILE: {os.path.basename(file_path)}")
            print(f"{'='*80}")
            
            # Try to load, repair if corrupted
            try:
                self.sf2_workbook = load_workbook(file_path)
            except KeyError as e:
                if "Content_Types" in str(e):
                    print(f"⚠️  File corrupted, attempting repair...")
                    if self.repair_excel_file(file_path):
                        print(f"✅ Repair successful!")
                        self.sf2_workbook = load_workbook(file_path)
                    else:
                        print(f"❌ Cannot repair")
                        messagebox.showerror("Error", "Cannot repair corrupted file")
                        return
                else:
                    raise
            
            self.sf2_sheet = self.sf2_workbook.active
            self.sf2_file = file_path
            
            self.student_names = []
            self.scanned_today = []
            self.existing_marks = {}  # Reset
            
            # SIMPLE LOGIC: Find date column, use that column!
            today = datetime.now()
            day_of_month = today.day
            
            print(f"\n🔍 Looking for date: {day_of_month}")
            print(f"Searching Row 11 for the date...")
            
            date_column = None
            for col in range(1, self.sf2_sheet.max_column + 1):
                cell_value = self.sf2_sheet.cell(11, col).value
                
                if cell_value is not None:
                    try:
                        date_num = int(cell_value)
                        if date_num == day_of_month:
                            date_column = col
                            print(f"✅ FOUND! Date {day_of_month} in Column {col}")
                            break
                    except (ValueError, TypeError):
                        continue
            
            if date_column is None:
                print(f"⚠️  Date {day_of_month} NOT FOUND in Row 11")
                self.date_status.config(text=f"📅 Date: {day_of_month} NOT FOUND", fg=self.RED)
                self.current_column = None
            else:
                # SIMPLE: Use that same column!
                day_letter = self.sf2_sheet.cell(12, date_column).value
                
                print(f"✅ Will mark attendance in Column {date_column} ({day_letter})")
                
                self.current_column = date_column
                self.date_status.config(
                    text=f"📅 Date: {day_of_month} ({day_letter}) → Col {date_column}",
                    fg=self.GREEN
                )
            
            # Load students
            print(f"\n👥 Loading students from Column B...")
            print("-" * 80)
            
            for row in range(13, self.sf2_sheet.max_row + 1):
                name_cell = self.sf2_sheet.cell(row, 2).value
                
                if not name_cell:
                    continue
                
                num_cell = self.sf2_sheet.cell(row, 1).value
                student_num = str(num_cell).strip() if num_cell else ""
                
                if self.is_valid_student_name(name_cell):
                    name = name_cell.strip()
                    self.student_names.append({
                        "name": name,
                        "number": student_num,
                        "row": row
                    })
                    
                    # ✨ CHECK EXISTING MARKS IN TODAY'S COLUMN!
                    if self.current_column:
                        existing_mark = self.sf2_sheet.cell(row, self.current_column).value
                        if existing_mark and str(existing_mark).strip() == "✓":
                            self.existing_marks[name] = True
                            print(f"  ✓ {student_num:3s} | {name} (already marked)")
                        else:
                            self.existing_marks[name] = False
                            print(f"    {student_num:3s} | {name}")
                    else:
                        self.existing_marks[name] = False
                        print(f"    {student_num:3s} | {name}")
            
            print("-" * 80)
            print(f"✅ Loaded {len(self.student_names)} students")
            print(f"{'='*80}\n")
            
            # Update UI
            self.file_status.config(
                text=f"📄 File: {os.path.basename(file_path)}",
                fg=self.GREEN
            )
            self.student_count_label.config(text=f"👥 Students: {len(self.student_names)}")
            self.total_label.config(text=str(len(self.student_names)))
            
            self.update_preview()
            self.update_counters()
        
        except Exception as e:
            print(f"❌ Error: {e}")
            traceback.print_exc()
            messagebox.showerror("Error", f"Failed to load file: {e}")
    
    def update_preview(self):
        """Update preview tab"""
        for item in self.preview_tree.get_children():
            self.preview_tree.delete(item)
        
        for student in self.student_names:
            # Check if has existing mark OR scanned in this session
            has_existing = self.existing_marks.get(student['name'], False)
            has_new_scan = any(s['name'] == student['name'] for s in self.scanned_today)
            
            if has_existing or has_new_scan:
                status = "✅ Present"
            else:
                status = "❌ Absent"
            
            self.preview_tree.insert("", "end", values=(
                student['number'],
                student['name'],
                status
            ))
    
    def update_counters(self):
        """Update counters - ACCURATE: existing + new scans"""
        # Count existing marks
        existing_present = len([x for x in self.existing_marks.values() if x])
        
        # Count new scans
        new_scans = len(self.scanned_today)
        
        # ACCURATE TOTAL
        present = existing_present + new_scans
        absent = len(self.student_names) - present
        total = len(self.student_names)
        
        self.present_label.config(text=str(present))
        self.absent_label.config(text=str(absent))
        self.total_label.config(text=str(total))
    
    def start_camera(self):
        """Start camera with proper threading"""
        if self.camera_active:
            return
        
        if not self.sf2_file:
            messagebox.showwarning("Warning", "Please load a file first!")
            return
        
        self.camera = cv2.VideoCapture(0)
        if not self.camera.isOpened():
            messagebox.showerror("Error", "Cannot open camera!")
            return
        
        # Set camera properties for faster capture
        self.camera.set(cv2.CAP_PROP_BUFFERSIZE, 1)
        self.camera.set(cv2.CAP_PROP_FPS, 30)
        
        self.camera_active = True
        self.start_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)
        
        # Start camera thread
        self.camera_thread = threading.Thread(target=self.camera_thread_worker, daemon=True)
        self.camera_thread.start()
        
        print("📷 Camera started (continuous)")
        self.update_camera_frame()
    
    def camera_thread_worker(self):
        """Worker thread for camera - reads frames continuously"""
        while self.camera_active and self.camera and self.camera.isOpened():
            ret, frame = self.camera.read()
            if ret:
                try:
                    # Only keep latest frame, discard old ones
                    if self.frame_queue.full():
                        try:
                            self.frame_queue.get_nowait()
                        except queue.Empty:
                            pass
                    self.frame_queue.put(frame, block=False)
                except queue.Full:
                    pass
            else:
                break
            time.sleep(0.01)  # Small delay to prevent CPU hogging
    
    def update_camera_frame(self):
        """Update camera frame - process from queue continuously"""
        if not self.camera_active:
            return
        
        frame = None
        try:
            # Get the latest frame from queue (non-blocking)
            frame = self.frame_queue.get(timeout=0.1)
        except queue.Empty:
            # No frame yet, try again soon
            if self.camera_active:
                self.root.after(10, self.update_camera_frame)
            return
        
        if frame is None:
            if self.camera_active:
                self.root.after(10, self.update_camera_frame)
            return
        
        # PROCESS FRAME FOR QR CODES
        try:
            decoded_objects = decode(frame)
            
            for obj in decoded_objects:
                try:
                    qr_data = obj.data.decode('utf-8').strip()
                    
                    if self.is_valid_student_name(qr_data):
                        matching_student = next((s for s in self.student_names if s['name'] == qr_data), None)
                        
                        if matching_student:
                            # CHECK: Already has ✓ from before?
                            has_existing_mark = self.existing_marks.get(qr_data, False)
                            
                            # CHECK: Already scanned in THIS session?
                            already_scanned = any(s['name'] == qr_data for s in self.scanned_today)
                            
                            # Prevent rapid re-scans within 1 second
                            current_time = datetime.now().timestamp()
                            rapid_rescan = (qr_data == self.last_scanned and 
                                          (current_time - self.last_scan_time) < 1)
                            
                            if has_existing_mark:
                                print(f"⚠️  {qr_data}: Already marked from before!")
                            elif already_scanned:
                                print(f"⚠️  Already scanned in this session: {qr_data}")
                            elif rapid_rescan:
                                pass  # Silently ignore rapid rescans
                            else:
                                # NEW SCAN!
                                self.scanned_today.append({
                                    'name': qr_data,
                                    'time': datetime.now().strftime("%H:%M:%S")
                                })
                                self.last_scanned = qr_data
                                self.last_scan_time = current_time
                                
                                print(f"✅ Scanned: {qr_data}")
                                self.update_student_list()
                                self.update_counters()
                                self.update_preview()
                                
                                # AUTO-SAVE!
                                self.auto_save_attendance()
                        else:
                            pass  # Silently ignore unknown QR
                except Exception as e:
                    pass
        except Exception as e:
            pass
        
        # DRAW QR BOXES ON FRAME
        try:
            decoded_objects = decode(frame)
            for obj in decoded_objects:
                points = obj.polygon
                if len(points) > 0:
                    pts = [(int(p.x), int(p.y)) for p in points]
                    cv2.polylines(frame, [pts], True, (0, 255, 0), 2)
        except Exception as e:
            pass
        
        # DISPLAY FRAME
        try:
            image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            image = cv2.resize(image, (500, 400))
            image = Image.fromarray(image)
            photo = ImageTk.PhotoImage(image)
            
            self.camera_label.config(image=photo)
            self.camera_label.image = photo
        except Exception as e:
            pass
        
        # Schedule next update ASAP
        if self.camera_active:
            self.root.after(10, self.update_camera_frame)
    
    def stop_camera(self):
        """Stop camera"""
        self.camera_active = False
        
        # Wait a bit for thread to finish
        if self.camera_thread:
            self.camera_thread.join(timeout=1.0)
        
        if self.camera:
            self.camera.release()
            self.camera = None
        
        self.start_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)
        
        print("⏹ Camera stopped")
        
        self.start_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)
        
        print("⏹ Camera stopped")
    
    def update_student_list(self):
        """Update scanned list"""
        for item in self.student_tree.get_children():
            self.student_tree.delete(item)
        
        for student in self.scanned_today:
            self.student_tree.insert("", "end", values=(student['name'], student['time']))
    
    def auto_save_attendance(self):
        """Auto-save attendance after each scan"""
        try:
            if not self.sf2_file or self.current_column is None:
                return
            
            # CHECK IF EXCEL IS OPEN!
            if self.is_excel_file_open(self.sf2_file):
                print(f"⚠️  WARNING: Excel file is OPEN! Cannot save!")
                messagebox.showwarning("Excel Open", 
                    f"❌ Excel file is currently open!\n\n"
                    f"Close the file in Excel before scanning more students.\n\n"
                    f"The system cannot write while Excel has the file locked!")
                return
            
            # Mark the last scanned student
            if self.scanned_today:
                last_scanned = self.scanned_today[-1]['name']
                
                for student in self.student_names:
                    if student['name'] == last_scanned:
                        row = student['row']
                        self.sf2_sheet.cell(row, self.current_column).value = "✓"
                        print(f"  💾 Auto-saved: {last_scanned}")
                        break
                
                # Save file
                self.sf2_workbook.save(self.sf2_file)
        
        except Exception as e:
            print(f"❌ Auto-save error: {e}")
    
    def refresh_file_list(self):
        """Refresh file list"""
        try:
            for item in self.file_tree.get_children():
                self.file_tree.delete(item)
            
            files = [f for f in os.listdir(self.active_folder) 
                    if f.endswith(('.xlsx', '.xls')) and not f.startswith('~')]
            
            for filename in sorted(files):
                filepath = os.path.join(self.active_folder, filename)
                size = os.path.getsize(filepath) / 1024
                modified = datetime.fromtimestamp(os.path.getmtime(filepath)).strftime("%Y-%m-%d %H:%M")
                
                self.file_tree.insert("", "end", values=(filename, f"{size:.1f}KB", modified))
        except Exception as e:
            print(f"Error: {e}")
    
    def browse_file(self):
        """Browse for file"""
        file_path = filedialog.askopenfilename(
            initialdir=self.active_folder,
            filetypes=[("Excel", "*.xlsx *.xls"), ("All", "*.*")]
        )
        
        if file_path:
            self.load_file(file_path)
    
    def open_qr_folder(self):
        """Open QR folder"""
        try:
            import subprocess
            if os.name == 'nt':
                subprocess.Popen(f'explorer "{self.qr_folder}"')
            else:
                subprocess.Popen(['open', self.qr_folder])
        except Exception as e:
            messagebox.showerror("Error", f"Cannot open folder: {e}")
    
    def open_active_folder(self):
        """Open Active folder"""
        try:
            import subprocess
            if os.name == 'nt':
                subprocess.Popen(f'explorer "{self.active_folder}"')
            else:
                subprocess.Popen(['open', self.active_folder])
        except Exception as e:
            messagebox.showerror("Error", f"Cannot open folder: {e}")
    
    def open_output_folder(self):
        """Open Output folder (where files are saved)"""
        try:
            import subprocess
            output_folder = os.path.join(os.path.expanduser("~"), "Downloads", "ATStudios-Project")
            if os.name == 'nt':
                subprocess.Popen(f'explorer "{output_folder}"')
            else:
                subprocess.Popen(['open', output_folder])
        except Exception as e:
            messagebox.showerror("Error", f"Cannot open folder: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = AttendanceSystem(root)
    root.mainloop()
