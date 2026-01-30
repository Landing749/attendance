"""
Dr. Alfredo Pio De Roda ES - QR Attendance System (Toga Version)
Version: PERFECTED TOGA - Exact UI Match from Tkinter
Date: January 30, 2026

PERFECTED FEATURES:
1. ✅ EXACT UI from Tkinter version (layout, colors, styling)
2. ✅ FILTER: Comprehensive student name validation
3. ✅ CHECK EXCEL: Existing marks on load
4. ✅ ACCURATE COUNTER: Existing + new scans
5. ✅ SMOOTH CAMERA: Threading for no freezing
6. ✅ AUTO-SAVE: Each scan saved immediately
"""

import toga
from toga.style import Pack
from toga.style.pack import COLUMN, ROW
import cv2
from PIL import Image
from pyzbar.pyzbar import decode
from openpyxl import load_workbook
from datetime import datetime
import os
import re
import threading
import queue
import time
import subprocess
from pathlib import Path


class AttendanceSystem(toga.App):
    def startup(self):
        """Setup the application"""
        # Initialize variables (avoid 'camera' - it's a reserved Toga property)
        self.video_capture = None
        self.camera_active = False
        self.camera_thread = None
        self.frame_queue = queue.Queue(maxsize=1)
        self.sf2_workbook = None
        self.sf2_sheet = None
        self.sf2_file = None
        self.student_names = []
        self.scanned_today = []
        self.existing_marks = {}  # Track existing ✓ from Excel
        self.current_column = None
        self.last_scanned = None  # Track last scan to prevent rapid re-scans
        self.last_scan_time = 0  # Track scan time
        self.temp_image_path = None  # Store persistent temp path
        
        # Dark theme colors (EXACT match to Tkinter)
        self.BG_DARK = "#0f1419"
        self.BG_CARD = "#1a202c"
        self.BG_INPUT = "#2d3748"
        self.BLUE = "#3b82f6"
        self.GREEN = "#10b981"
        self.RED = "#ef4444"
        self.YELLOW = "#f59e0b"
        self.CYAN = "#06b6d4"
        self.PURPLE = "#8b5cf6"
        self.TEXT_PRIMARY = "#ffffff"
        self.TEXT_SECONDARY = "#9ca3af"
        
        # Setup folders
        self.home_dir = Path.home()
        self.base_folder = self.home_dir / "SF2_Files"
        self.active_folder = self.base_folder / "Active"
        self.backup_folder = self.base_folder / "Backups"
        self.archive_folder = self.base_folder / "Archive"
        self.qr_folder = self.base_folder / "QR_Codes"
        
        for folder in [self.active_folder, self.backup_folder, self.archive_folder, self.qr_folder]:
            folder.mkdir(parents=True, exist_ok=True)
        
        # Create persistent temp image path
        self.temp_image_path = self.home_dir / "camera_feed.jpg"
        
        # Build UI
        self.main_window = toga.MainWindow(title="QR Attendance System - Dr. Alfredo Pio De Roda ES")
        
        # Create tab container
        self.setup_ui()
        
        # Auto-load file
        self.auto_load_file()
        
        self.main_window.show()
    
    def is_valid_student_name(self, name):
        """Validate if text is a real student name with comprehensive filtering"""
        if not name or not isinstance(name, str):
            return False
        
        name = name.strip()
        if len(name) < 2:
            return False
        
        name_upper = name.upper()
        
        # COMPREHENSIVE EXCLUSION LIST (same as Tkinter version)
        excluded_patterns = [
            "SUMIF", "COUNTIF", "AVERAGE", "SUM(", "COUNT(", "IF(",
            "VLOOKUP", "HLOOKUP", "INDEX", "MATCH",
            "SCHOOL FORM", "SF2", "DAILY ATTENDANCE", "ATTENDANCE REPORT",
            "LEARNER'S NAME", "LAST NAME", "FIRST NAME", "MIDDLE NAME",
            "CODES FOR CHECKING", "PRESENT", "ABSENT", "TARDY",
            "HALF SHADED", "UPPER", "LOWER", "CUTTING CLASSES", "LATE COMER",
            "DROPPED", "TRANSFERRED", "ENROLLED", "REGISTRATION",
            "TOTAL", "COMBINED", "PER DAY", "SUMMARY", "MALE", "FEMALE",
            "MONTH:", "BLANK", "(BLANK)", "NO. OF DAYS", "CLASSES",
            "PERCENTAGE", "ENROLMENT", "AVERAGE DAILY", "ATTENDANCE",
            "REGISTERED LEARNERS", "END OF THE MONTH", "SCHOOL YEAR",
            "1ST FRIDAY", "REPORTING MONTH", "SCHOOL DAYS",
            "REASONS", "CAUSES", "DROPPING OUT", "DROP OUT", "DROPOUT",
            "DOMESTIC-RELATED", "INDIVIDUAL-RELATED", "SCHOOL-RELATED",
            "GEOGRAPHIC", "ENVIRONMENTAL", "FINANCIAL-RELATED",
            "TAKE CARE", "SIBLINGS", "EARLY MARRIAGE", "PREGNANCY",
            "PARENTS' ATTITUDE", "FAMILY PROBLEMS", "ILLNESS",
            "OVERAGE", "DEATH", "DRUG ABUSE", "ACADEMIC PERFORMANCE",
            "LACK OF INTEREST", "DISTRACTIONS", "HUNGER", "MALNUTRITION",
            "TEACHER FACTOR", "PHYSICAL CONDITION", "CLASSROOM",
            "PEER INFLUENCE", "DISTANCE", "HOME AND SCHOOL",
            "ARMED CONFLICT", "TRIBAL WARS", "CLAN FEUDS",
            "CALAMITIES", "DISASTERS", "CHILD LABOR", "WORK",
            "OTHERS (SPECIFY)",
            "GUIDELINES:", "ACCOMPLISHED", "REFER", "DATES SHALL",
            "WRITTEN IN", "COLUMNS AFTER", "COMPUTE", "FOLLOWING",
            "EVERY END", "ADVISER", "SUBMIT", "OFFICE", "PRINCIPAL",
            "RECORDING", "SUMMARY TABLE", "FORM 4", "SIGNED",
            "RETURNED", "PROVIDE", "NECESSARY", "INTERVENTIONS",
            "HOME VISITATION", "ABSENT FOR 5", "CONSECUTIVE DAYS",
            "RISK OF", "PERFORMANCE", "REFLECTED", "FORM 137", "FORM 138",
            "GRADING PERIOD", "BEGINNING", "CUT-OFF",
            "MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY",
            "MONDAY,", "TUESDAY,", "WEDNESDAY,", "THURSDAY,", "FRIDAY,",
            "CERTIFY", "TRUE", "CORRECT", "REPORT", "SIGNATURE",
            "PRINTED NAME", "TEACHER", "SCHOOL HEAD", "ATTESTED",
            "PAGE", "OF", "SCHOOL FORM 2", "___",
            "LEARNER", "STUDENT", "NAME", "NAMES", "ID", "NUMBER",
            "ENROLLMENT", "ENROL",
            "NAN", "NONE", "N/A", "NULL", "BLANK", "EMPTY",
            "PERCENTAGE OF ENROLMENT", "PERCENTAGE OF ENROLLMENT",
            "AVERAGE DAILY ATTENDANCE", 
            "PERCENTAGE OF ATTENDANCE FOR THE MONTH",
            "PERCENTAGE OF ATTENDANCE",
        ]
        
        for pattern in excluded_patterns:
            if pattern in name_upper:
                return False
        
        if not any(c.isalpha() for c in name):
            return False
        
        if name.replace('.', '').replace(',', '').replace(' ', '').isdigit():
            return False
        
        if re.match(r'^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}$', name):
            return False
        
        if not any(c.isalnum() for c in name):
            return False
        
        return True
    
    def is_excel_file_open(self, file_path):
        """Check if Excel file is open/locked"""
        try:
            temp_name = str(file_path) + ".tmp"
            os.rename(file_path, temp_name)
            os.rename(temp_name, file_path)
            return False
        except (OSError, IOError):
            return True
    
    def setup_ui(self):
        """Setup the complete UI with EXACT Tkinter layout"""
        self.scan_tab = self.setup_scan_tab()
        self.files_tab = self.setup_files_tab()
        self.preview_tab = self.setup_preview_tab()
        self.settings_tab = self.setup_settings_tab()
        
        option_container = toga.OptionContainer(
            content=[
                ("📱 SCAN", self.scan_tab),
                ("📂 FILES", self.files_tab),
                ("📊 PREVIEW", self.preview_tab),
                ("⚙️ SETTINGS", self.settings_tab),
            ],
            style=Pack(flex=1)
        )
        
        self.main_window.content = option_container
    
    def setup_scan_tab(self):
        """Create SCAN tab - EXACT Tkinter layout"""
        main_box = toga.Box(style=Pack(direction=COLUMN, padding=10))
        
        # TOP SECTION: Camera + System Info (side by side)
        top_container = toga.Box(style=Pack(direction=ROW, padding=5, flex=1))
        
        # ===== LEFT SIDE: CAMERA =====
        left_box = toga.Box(style=Pack(direction=COLUMN, padding=10, flex=1))
        
        # Camera header
        camera_header = toga.Label(
            "📷 QR Scanner",
            style=Pack(padding=(0, 5), font_size=14, font_weight='bold')
        )
        left_box.add(camera_header)
        
        # Camera display - FIXED: Initialize with empty image to ensure it renders
        self.camera_label = toga.ImageView(
            style=Pack(width=640, height=480, padding=5)
        )
        left_box.add(self.camera_label)
        
        # Camera controls
        camera_controls = toga.Box(style=Pack(direction=ROW, padding=5))
        self.start_btn = toga.Button(
            "▶ START SCANNING",
            on_press=self.start_camera,
            style=Pack(flex=1, padding=2)
        )
        self.stop_btn = toga.Button(
            "⏹ STOP SCANNING",
            on_press=self.stop_camera,
            enabled=False,
            style=Pack(flex=1, padding=2)
        )
        camera_controls.add(self.start_btn)
        camera_controls.add(self.stop_btn)
        left_box.add(camera_controls)
        
        top_container.add(left_box)
        
        # ===== RIGHT SIDE: SYSTEM INFO =====
        right_box = toga.Box(style=Pack(direction=COLUMN, padding=10, flex=1))
        
        # System Status section
        status_header = toga.Label(
            "📊 System Status",
            style=Pack(padding=(0, 5), font_size=14, font_weight='bold')
        )
        right_box.add(status_header)
        
        self.file_status = toga.Label(
            "📁 File: Not loaded",
            style=Pack(padding=2)
        )
        self.date_status = toga.Label(
            "📅 Date: --",
            style=Pack(padding=2)
        )
        self.students_status = toga.Label(
            "👥 Students: 0",
            style=Pack(padding=2)
        )
        right_box.add(self.file_status)
        right_box.add(self.date_status)
        right_box.add(self.students_status)
        
        right_box.add(toga.Divider(style=Pack(padding=5)))
        
        # Today's Attendance section
        attendance_header = toga.Label(
            "📈 Today's Attendance",
            style=Pack(padding=(0, 5), font_size=14, font_weight='bold')
        )
        right_box.add(attendance_header)
        
        self.present_label = toga.Label(
            "✅ Present: 0",
            style=Pack(padding=2)
        )
        self.absent_label = toga.Label(
            "❌ Absent: 0",
            style=Pack(padding=2)
        )
        self.total_label = toga.Label(
            "📊 Total: 0",
            style=Pack(padding=2)
        )
        right_box.add(self.present_label)
        right_box.add(self.absent_label)
        right_box.add(self.total_label)
        
        right_box.add(toga.Divider(style=Pack(padding=5)))
        
        # Scanned Students section
        scanned_header = toga.Label(
            "✅ Scanned Students (Auto-Saved)",
            style=Pack(padding=(0, 5), font_size=14, font_weight='bold')
        )
        right_box.add(scanned_header)
        
        self.student_tree = toga.Table(
            headings=["Student Name", "Time"],
            data=[],
            accessors=["name", "time"],
            style=Pack(flex=1, padding=5)
        )
        right_box.add(self.student_tree)
        
        top_container.add(right_box)
        main_box.add(top_container)
        
        # BOTTOM SECTION: Action buttons (NO SAVE BUTTON - IT'S AUTO)
        actions_box = toga.Box(style=Pack(direction=ROW, padding=10))
        
        qr_folder_btn = toga.Button(
            "📂 OPEN QR FOLDER",
            on_press=self.open_qr_folder,
            style=Pack(flex=1, padding=5)
        )
        active_folder_btn = toga.Button(
            "📊 OPEN ACTIVE FOLDER",
            on_press=self.open_active_folder,
            style=Pack(flex=1, padding=5)
        )
        
        actions_box.add(qr_folder_btn)
        actions_box.add(active_folder_btn)
        
        main_box.add(actions_box)
        
        return main_box
    
    def setup_files_tab(self):
        """Create FILES tab - EXACT Tkinter layout"""
        main_box = toga.Box(style=Pack(direction=COLUMN, padding=10))
        
        # Header
        header = toga.Label(
            "📂 File Management",
            style=Pack(padding=(5, 10), font_size=16, font_weight='bold')
        )
        main_box.add(header)
        
        # Current file info
        current_box = toga.Box(style=Pack(direction=COLUMN, padding=10))
        current_label = toga.Label(
            "📁 Current File:",
            style=Pack(padding=2, font_weight='bold')
        )
        self.current_file_label = toga.Label(
            "Not loaded",
            style=Pack(padding=2)
        )
        current_box.add(current_label)
        current_box.add(self.current_file_label)
        main_box.add(current_box)
        
        main_box.add(toga.Divider(style=Pack(padding=10)))
        
        # Available files section
        list_header = toga.Label(
            "📋 Available Files (Active Folder):",
            style=Pack(padding=(5, 5), font_weight='bold')
        )
        main_box.add(list_header)
        
        self.file_tree = toga.Table(
            headings=["Filename", "Size", "Modified"],
            data=[],
            accessors=["filename", "size", "modified"],
            style=Pack(flex=1, padding=5)
        )
        main_box.add(self.file_tree)
        
        # Action buttons
        actions_box = toga.Box(style=Pack(direction=ROW, padding=10))
        
        refresh_btn = toga.Button(
            "🔄 REFRESH LIST",
            on_press=self.refresh_file_list,
            style=Pack(flex=1, padding=5)
        )
        browse_btn = toga.Button(
            "📂 BROWSE FILE",
            on_press=self.browse_file,
            style=Pack(flex=1, padding=5)
        )
        
        actions_box.add(refresh_btn)
        actions_box.add(browse_btn)
        main_box.add(actions_box)
        
        return main_box
    
    def setup_preview_tab(self):
        """Create PREVIEW tab - EXACT Tkinter layout"""
        main_box = toga.Box(style=Pack(direction=COLUMN, padding=10))
        
        # Header
        header = toga.Label(
            "📊 Student Attendance Preview",
            style=Pack(padding=(5, 10), font_size=16, font_weight='bold')
        )
        main_box.add(header)
        
        # Preview table
        self.preview_tree = toga.Table(
            headings=["No.", "Student Name", "Status"],
            data=[],
            accessors=["number", "name", "status"],
            style=Pack(flex=1, padding=5)
        )
        main_box.add(self.preview_tree)
        
        # Action buttons
        actions_box = toga.Box(style=Pack(direction=ROW, padding=10))
        
        refresh_btn = toga.Button(
            "🔄 REFRESH PREVIEW",
            on_press=self.update_preview,
            style=Pack(flex=1, padding=5)
        )
        
        actions_box.add(refresh_btn)
        main_box.add(actions_box)
        
        return main_box
    
    def setup_settings_tab(self):
        """Create SETTINGS tab - EXACT Tkinter layout"""
        main_box = toga.Box(style=Pack(direction=COLUMN, padding=10))
        
        # Header
        header = toga.Label(
            "⚙️ System Settings & Information",
            style=Pack(padding=(5, 10), font_size=16, font_weight='bold')
        )
        main_box.add(header)
        
        # System info
        info_box = toga.Box(style=Pack(direction=COLUMN, padding=10))
        
        app_label = toga.Label(
            "📱 Application: QR Attendance System v1.0 PERFECTED",
            style=Pack(padding=3)
        )
        school_label = toga.Label(
            "🏫 School: Dr. Alfredo Pio De Roda ES",
            style=Pack(padding=3)
        )
        info_box.add(app_label)
        info_box.add(school_label)
        
        info_box.add(toga.Divider(style=Pack(padding=10)))
        
        folder_header = toga.Label(
            "📂 Folder Locations:",
            style=Pack(padding=3, font_weight='bold')
        )
        info_box.add(folder_header)
        
        base_label = toga.Label(
            f"Base Folder: {self.base_folder}",
            style=Pack(padding=2)
        )
        active_label = toga.Label(
            f"Active Folder: {self.active_folder}",
            style=Pack(padding=2)
        )
        backup_label = toga.Label(
            f"Backup Folder: {self.backup_folder}",
            style=Pack(padding=2)
        )
        archive_label = toga.Label(
            f"Archive Folder: {self.archive_folder}",
            style=Pack(padding=2)
        )
        qr_label = toga.Label(
            f"QR Codes Folder: {self.qr_folder}",
            style=Pack(padding=2)
        )
        
        info_box.add(base_label)
        info_box.add(active_label)
        info_box.add(backup_label)
        info_box.add(archive_label)
        info_box.add(qr_label)
        
        main_box.add(info_box)
        
        main_box.add(toga.Divider(style=Pack(padding=10)))
        
        # Action buttons
        actions_box = toga.Box(style=Pack(direction=ROW, padding=10))
        
        qr_btn = toga.Button(
            "📂 OPEN QR FOLDER",
            on_press=self.open_qr_folder,
            style=Pack(flex=1, padding=5)
        )
        active_btn = toga.Button(
            "📊 OPEN ACTIVE FOLDER",
            on_press=self.open_active_folder,
            style=Pack(flex=1, padding=5)
        )
        
        actions_box.add(qr_btn)
        actions_box.add(active_btn)
        main_box.add(actions_box)
        
        return main_box
    
    def auto_load_file(self):
        """Auto-load the most recent file"""
        try:
            files = list(self.active_folder.glob("*.xlsx"))
            files = [f for f in files if not f.name.startswith('~')]
            
            if files:
                most_recent = max(files, key=lambda f: f.stat().st_mtime)
                self.load_file(most_recent)
        except Exception as e:
            print(f"Auto-load error: {e}")
    
    def load_file(self, file_path):
        """Load SF2 file with EXACT Tkinter logic"""
        try:
            if isinstance(file_path, list):
                file_path = file_path[0] if file_path else None
            
            if not file_path:
                return
            
            file_path = Path(file_path)
            
            print(f"\n{'='*80}")
            print(f"LOADING FILE: {file_path.name}")
            print(f"{'='*80}")
            
            # Check if Excel is open
            if self.is_excel_file_open(file_path):
                self.main_window.error_dialog(
                    "Excel File Open",
                    "❌ Excel file is currently open!\n\nClose the file in Excel before loading."
                )
                return
            
            # Load workbook
            self.sf2_workbook = load_workbook(file_path)
            self.sf2_sheet = self.sf2_workbook.active
            self.sf2_file = file_path
            
            self.student_names = []
            self.scanned_today = []
            self.existing_marks = {}  # Reset
            
            # EXACT TKINTER LOGIC: Find date in Row 11
            today = datetime.now()
            day_of_month = today.day
            
            print(f"\n🔍 Looking for date: {day_of_month}")
            print(f"Searching Row 11 for the date...")
            
            date_column = None
            for col in range(1, self.sf2_sheet.max_column + 1):
                cell_value = self.sf2_sheet.cell(11, col).value
                
                if cell_value is not None:
                    try:
                        date_num = int(cell_value)
                        if date_num == day_of_month:
                            date_column = col
                            print(f"✅ FOUND! Date {day_of_month} in Column {col}")
                            break
                    except (ValueError, TypeError):
                        continue
            
            if date_column is None:
                print(f"⚠️  Date {day_of_month} NOT FOUND in Row 11")
                self.date_status.text = f"📅 Date: {day_of_month} NOT FOUND"
                self.current_column = None
            else:
                # Get day letter from Row 12
                day_letter = self.sf2_sheet.cell(12, date_column).value
                
                print(f"✅ Will mark attendance in Column {date_column} ({day_letter})")
                
                self.current_column = date_column
                self.date_status.text = f"📅 Date: {day_of_month} ({day_letter}) → Col {date_column}"
            
            # EXACT TKINTER LOGIC: Load students from Column B (column 2), starting Row 13
            print(f"\n👥 Loading students from Column B...")
            print("-" * 80)
            
            for row in range(13, self.sf2_sheet.max_row + 1):
                name_cell = self.sf2_sheet.cell(row, 2).value  # Column B
                
                if not name_cell:
                    continue
                
                num_cell = self.sf2_sheet.cell(row, 1).value  # Column A
                student_num = str(num_cell).strip() if num_cell else ""
                
                if self.is_valid_student_name(name_cell):
                    name = name_cell.strip()
                    self.student_names.append({
                        "name": name,
                        "number": student_num,
                        "row": row
                    })
                    
                    # CHECK EXISTING MARKS IN TODAY'S COLUMN!
                    if self.current_column:
                        existing_mark = self.sf2_sheet.cell(row, self.current_column).value
                        if existing_mark and str(existing_mark).strip() == "✓":
                            self.existing_marks[name] = True
                            print(f"  ✓ {student_num:3s} | {name} (already marked)")
                        else:
                            self.existing_marks[name] = False
                            print(f"    {student_num:3s} | {name}")
                    else:
                        self.existing_marks[name] = False
                        print(f"    {student_num:3s} | {name}")
            
            print("-" * 80)
            print(f"✅ Loaded {len(self.student_names)} students")
            print(f"{'='*80}\n")
            
            # Update UI
            self.file_status.text = f"📁 File: {file_path.name}"
            self.students_status.text = f"👥 Students: {len(self.student_names)}"
            self.current_file_label.text = file_path.name
            
            # Update counters and preview
            self.update_student_list()
            self.update_counters()
            self.update_preview(None)
            
        except Exception as e:
            print(f"❌ Load error: {e}")
            import traceback
            traceback.print_exc()
            self.main_window.error_dialog("Error", f"Failed to load file:\n{e}")
    
    def start_camera(self, widget):
        """Start camera with EXACT Tkinter logic - MOBILE OPTIMIZED"""
        try:
            # Try to open camera (index 0 for mobile, or auto-detect)
            print("📷 Attempting to open camera...")
            
            # For Windows: Try DirectShow backend first (more reliable)
            if os.name == 'nt':
                print("   Using DirectShow backend for Windows...")
                self.video_capture = cv2.VideoCapture(0, cv2.CAP_DSHOW)
            else:
                self.video_capture = cv2.VideoCapture(0)
            
            # Wait a moment for camera to initialize
            time.sleep(0.5)
            
            # If camera 0 fails, try different indices (for mobile devices)
            if not self.video_capture.isOpened():
                print("⚠️  Camera 0 failed, trying other indices...")
                for i in range(1, 5):
                    print(f"   Trying camera index {i}...")
                    if os.name == 'nt':
                        self.video_capture = cv2.VideoCapture(i, cv2.CAP_DSHOW)
                    else:
                        self.video_capture = cv2.VideoCapture(i)
                    time.sleep(0.3)
                    if self.video_capture.isOpened():
                        print(f"✅ Camera {i} opened!")
                        break
            
            if not self.video_capture.isOpened():
                self.main_window.error_dialog(
                    "Camera Error", 
                    "Cannot access camera!\n\n"
                    "Possible issues:\n"
                    "• Camera is being used by another app\n"
                    "• Camera permissions not granted\n"
                    "• No camera detected\n\n"
                    "Try closing other apps using the camera."
                )
                return
            
            # Set camera properties for better mobile performance
            self.video_capture.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
            self.video_capture.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
            self.video_capture.set(cv2.CAP_PROP_FPS, 30)
            self.video_capture.set(cv2.CAP_PROP_BUFFERSIZE, 1)  # Reduce buffer for less lag
            
            # Test read
            ret, test_frame = self.video_capture.read()
            if not ret:
                self.video_capture.release()
                self.main_window.error_dialog(
                    "Camera Error",
                    "Camera opened but cannot read frames!\n\n"
                    "The camera may be locked by another application.\n"
                    "Please close other camera apps and try again."
                )
                return
            
            print(f"✅ Successfully read test frame: {test_frame.shape}")
            
            self.camera_active = True
            self.start_btn.enabled = False
            self.stop_btn.enabled = True
            
            # Start background thread
            self.camera_thread = threading.Thread(target=self.camera_worker, daemon=True)
            self.camera_thread.start()
            
            # Start UI updates using asyncio
            import asyncio
            asyncio.create_task(self.update_camera_loop())
            
            print("▶ Camera started successfully!")
        except Exception as e:
            print(f"❌ Camera error: {e}")
            import traceback
            traceback.print_exc()
            self.main_window.error_dialog("Error", f"Camera error: {e}")
    
    def camera_worker(self):
        """Background worker to read camera frames continuously for LIVE FEED"""
        consecutive_failures = 0
        max_failures = 30  # Stop after 30 consecutive failures
        frame_drop_counter = 0
        
        while self.camera_active:
            try:
                if not self.video_capture or not self.video_capture.isOpened():
                    print("⚠️  Camera not opened in worker thread")
                    break
                
                ret, frame = self.video_capture.read()
                
                if ret and frame is not None:
                    consecutive_failures = 0  # Reset counter on success
                    
                    # Drop oldest frame if queue is full to keep live feed smooth
                    if self.frame_queue.full():
                        try:
                            self.frame_queue.get_nowait()  # Remove oldest
                            frame_drop_counter += 1
                        except queue.Empty:
                            pass
                    
                    try:
                        self.frame_queue.put_nowait(frame)
                    except queue.Full:
                        pass  # Silently drop if still full
                else:
                    consecutive_failures += 1
                    if consecutive_failures >= max_failures:
                        print(f"❌ Camera worker: {max_failures} consecutive failures, stopping...")
                        break
                    time.sleep(0.05)  # Brief wait before retry
                    continue
                
                # Minimal sleep to maintain responsiveness without blocking
                time.sleep(0.005)
            except Exception as e:
                print(f"Camera worker error: {e}")
                consecutive_failures += 1
                if consecutive_failures >= max_failures:
                    break
                time.sleep(0.1)
        
        print(f"Camera worker thread stopped (dropped {frame_drop_counter} frames)")
    
    async def update_camera_loop(self):
        """Async loop to update camera display - OPTIMIZED FOR LIVE FEED"""
        import asyncio
        frame_counter = 0
        while self.camera_active:
            try:
                # Call update more frequently for smoother playback
                frame_counter += 1
                self.update_camera_frame()
                
                # Wait ~16ms for ~60 FPS update rate
                await asyncio.sleep(0.016)
            except Exception as e:
                print(f"Loop error: {e}")
                break
    
    def update_camera_frame(self):
        """Update camera display with EXACT Tkinter logic - OPTIMIZED FOR LIVE FEED"""
        try:
            frame = self.frame_queue.get_nowait()
        except queue.Empty:
            return
        
        # SCAN QR CODES
        try:
            decoded_objects = decode(frame)
            for obj in decoded_objects:
                qr_data = obj.data.decode('utf-8').strip()
                
                if self.is_valid_student_name(qr_data):
                    matching_student = any(s['name'] == qr_data for s in self.student_names)
                    
                    if matching_student:
                        # CHECK: Already has ✓ from before?
                        has_existing_mark = self.existing_marks.get(qr_data, False)
                        
                        # CHECK: Already scanned in THIS session?
                        already_scanned = any(s['name'] == qr_data for s in self.scanned_today)
                        
                        # Prevent rapid re-scans within 1 second
                        current_time = datetime.now().timestamp()
                        rapid_rescan = (qr_data == self.last_scanned and 
                                      (current_time - self.last_scan_time) < 1)
                        
                        if has_existing_mark:
                            print(f"⚠️  {qr_data}: Already marked from before!")
                        elif already_scanned:
                            print(f"⚠️  Already scanned in this session: {qr_data}")
                        elif rapid_rescan:
                            pass  # Silently ignore rapid rescans
                        else:
                            # NEW SCAN!
                            self.scanned_today.append({
                                'name': qr_data,
                                'time': datetime.now().strftime("%H:%M:%S")
                            })
                            self.last_scanned = qr_data
                            self.last_scan_time = current_time
                            
                            print(f"✅ Scanned: {qr_data}")
                            self.update_student_list()
                            self.update_counters()
                            self.update_preview(None)
                            
                            # AUTO-SAVE!
                            self.auto_save_attendance()
        except Exception as e:
            pass  # Silently ignore QR decode errors
        
        # DRAW QR BOXES
        try:
            decoded_objects = decode(frame)
            for obj in decoded_objects:
                points = obj.polygon
                if len(points) > 0:
                    pts = [(int(p.x), int(p.y)) for p in points]
                    import numpy as np
                    pts_array = np.array([pts], dtype=np.int32)
                    cv2.polylines(frame, pts_array, True, (0, 255, 0), 3)  # Thicker line for mobile
        except Exception as e:
            pass
        
        # DISPLAY FRAME - OPTIMIZED: Force image reload each frame for smooth live feed
        try:
            # Convert BGR to RGB
            image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            # Resize to fit display
            image = cv2.resize(image, (640, 480))
            pil_image = Image.fromarray(image)
            
            # Save with timestamp to force reload (prevents image caching)
            import time as time_module
            cache_bust = str(int(time_module.time() * 1000))
            temp_path = str(self.temp_image_path).replace('.jpg', f'_{cache_bust}.jpg')
            pil_image.save(temp_path, 'JPEG', quality=85)
            
            # Update image view with unique path to force reload
            self.camera_label.image = toga.Image(temp_path)
            
        except Exception as e:
            pass  # Silently continue on display errors
    
    def stop_camera(self, widget):
        """Stop camera with EXACT Tkinter logic"""
        self.camera_active = False
        
        if self.camera_thread:
            self.camera_thread.join(timeout=1.0)
        
        if self.video_capture:
            self.video_capture.release()
            self.video_capture = None
        
        self.start_btn.enabled = True
        self.stop_btn.enabled = False
        
        print("⏹ Camera stopped")
    
    def update_student_list(self):
        """Update scanned students table"""
        data = [{'name': s['name'], 'time': s['time']} for s in self.scanned_today]
        self.student_tree.data = data
    
    def update_counters(self):
        """Update attendance counters with EXACT Tkinter logic"""
        existing_present = sum(1 for v in self.existing_marks.values() if v)
        new_present = len(self.scanned_today)
        total_present = existing_present + new_present
        total_students = len(self.student_names)
        absent = total_students - total_present
        
        self.present_label.text = f"✅ Present: {total_present} (Existing: {existing_present} + New: {new_present})"
        self.absent_label.text = f"❌ Absent: {absent}"
        self.total_label.text = f"📊 Total: {total_students}"
    
    def update_preview(self, widget):
        """Update preview table with EXACT Tkinter logic"""
        if not self.student_names:
            return
        
        data = []
        for idx, student in enumerate(self.student_names, 1):
            name = student['name']
            
            if self.existing_marks.get(name, False):
                status = "✅ (Before)"
            elif any(s['name'] == name for s in self.scanned_today):
                status = "✅ (Today)"
            else:
                status = "⭕ Absent"
            
            data.append({
                'number': str(idx),
                'name': name,
                'status': status
            })
        
        self.preview_tree.data = data
    
    def auto_save_attendance(self):
        """Auto-save attendance after each scan with EXACT Tkinter logic"""
        try:
            if not self.sf2_file or self.current_column is None:
                return
            
            # CHECK IF EXCEL IS OPEN!
            if self.is_excel_file_open(self.sf2_file):
                print(f"⚠️  WARNING: Excel file is OPEN! Cannot save!")
                self.main_window.error_dialog(
                    "Excel Open",
                    "❌ Excel file is currently open!\n\n"
                    "Close the file in Excel before scanning more students.\n\n"
                    "The system cannot write while Excel has the file locked!"
                )
                return
            
            # Mark the last scanned student
            if self.scanned_today:
                last_scanned = self.scanned_today[-1]['name']
                
                for student in self.student_names:
                    if student['name'] == last_scanned:
                        row = student['row']
                        self.sf2_sheet.cell(row, self.current_column).value = "✓"
                        print(f"  💾 Auto-saved: {last_scanned}")
                        break
                
                # Save file
                self.sf2_workbook.save(self.sf2_file)
        
        except Exception as e:
            print(f"❌ Auto-save error: {e}")
    
    def refresh_file_list(self, widget):
        """Refresh file list"""
        try:
            files = list(self.active_folder.glob("*.xlsx"))
            files = [f for f in files if not f.name.startswith('~')]
            
            data = []
            for f in sorted(files, key=lambda x: x.stat().st_mtime, reverse=True):
                size_kb = f.stat().st_size / 1024
                modified = datetime.fromtimestamp(f.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
                data.append({
                    'filename': f.name,
                    'size': f"{size_kb:.1f} KB",
                    'modified': modified
                })
            
            self.file_tree.data = data
        except Exception as e:
            print(f"Error: {e}")
    
    def browse_file(self, widget):
        """Browse for file"""
        try:
            self.main_window.open_file_dialog(
                title="Select SF2 Excel File",
                initial_directory=self.active_folder,
                file_types=['xlsx', 'xls'],
                on_result=self.load_file
            )
        except Exception as e:
            print(f"Browse error: {e}")
    
    def open_qr_folder(self, widget):
        """Open QR folder"""
        try:
            if os.name == 'nt':
                subprocess.Popen(f'explorer "{self.qr_folder}"')
            else:
                subprocess.Popen(['open', str(self.qr_folder)])
        except Exception as e:
            self.main_window.error_dialog("Error", f"Cannot open folder: {e}")
    
    def open_active_folder(self, widget):
        """Open Active folder"""
        try:
            if os.name == 'nt':
                subprocess.Popen(f'explorer "{self.active_folder}"')
            else:
                subprocess.Popen(['open', str(self.active_folder)])
        except Exception as e:
            self.main_window.error_dialog("Error", f"Cannot open folder: {e}")


def main():
    return AttendanceSystem('QR Attendance System', 'org.dralfredroda.attendance')


if __name__ == '__main__':
    main().main_loop()
